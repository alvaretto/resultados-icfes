#!/usr/bin/env python3
"""
Script de prueba para verificar que todos los cambios se han aplicado correctamente.
Ejecutar con: python3 test_cambios.py
"""

import pandas as pd
import openpyxl
import sys

def test_excel_flexible():
    """Verifica que el archivo Excel de Modelo Flexible tiene la estructura correcta."""
    print("\n" + "="*80)
    print("TEST 1: Verificar estructura del archivo Excel - Modelo Flexible")
    print("="*80)
    
    file_path = "PCIELO-RESULTADOS-ICFES-MODELO-FLEXIBLE-2025.xlsx"
    
    try:
        # Cargar con openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Verificar fila 64 (2024)
        puntaje_2024 = ws.cell(row=64, column=13).value
        
        if puntaje_2024 != 203:
            print(f"❌ FALLO: Puntaje global 2024 es {puntaje_2024}, esperado 203")
            return False
        
        # Verificar que áreas de 2024 están vacías
        areas_vacias = all(ws.cell(row=64, column=col).value is None for col in range(8, 13))
        
        if not areas_vacias:
            print("❌ FALLO: Las áreas de 2024 no están vacías")
            return False
        
        print("✅ PASÓ: Estructura Excel correcta")
        print(f"   - Puntaje global 2024: {puntaje_2024}")
        print(f"   - Áreas 2024: Vacías (pendientes)")
        return True
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def test_pandas_indices():
    """Verifica que pandas lee los datos con los índices correctos."""
    print("\n" + "="*80)
    print("TEST 2: Verificar índices de pandas")
    print("="*80)
    
    file_path = "PCIELO-RESULTADOS-ICFES-MODELO-FLEXIBLE-2025.xlsx"
    
    try:
        df = pd.read_excel(file_path)
        
        # Verificar índice 62 (fila 64 Excel - 2024)
        puntaje_2024 = df.iloc[62]['Puntaje Global']
        
        if puntaje_2024 != 203:
            print(f"❌ FALLO: Puntaje global 2024 (índice 62) es {puntaje_2024}, esperado 203")
            return False
        
        # Verificar que áreas de 2024 están NaN
        areas_nan = all(pd.isna(df.iloc[62][area]) for area in 
                       ['Lectura Crítica', 'Matemáticas', 'Sociales y Ciudadanas', 
                        'Ciencias Naturales', 'Inglés'])
        
        if not areas_nan:
            print("❌ FALLO: Las áreas de 2024 no están NaN")
            return False
        
        # Verificar índice 61 (fila 63 Excel - 2025)
        puntaje_2025 = df.iloc[61]['Puntaje Global']
        
        if abs(puntaje_2025 - 213.54) > 0.01:
            print(f"❌ FALLO: Puntaje global 2025 es {puntaje_2025}, esperado ~213.54")
            return False
        
        print("✅ PASÓ: Índices de pandas correctos")
        print(f"   - Puntaje global 2024 (índice 62): {puntaje_2024}")
        print(f"   - Puntaje global 2025 (índice 61): {puntaje_2025:.2f}")
        print(f"   - Áreas 2024: NaN (pendientes)")
        return True
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def test_codigo_python():
    """Verifica que el código Python tiene sintaxis correcta."""
    print("\n" + "="*80)
    print("TEST 3: Verificar sintaxis del código Python")
    print("="*80)
    
    try:
        import py_compile
        py_compile.compile('app_resultados_icfes.py', doraise=True)
        print("✅ PASÓ: Sintaxis correcta en app_resultados_icfes.py")
        return True
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def test_aula_regular_intacta():
    """Verifica que el archivo de Aula Regular no fue modificado."""
    print("\n" + "="*80)
    print("TEST 4: Verificar que Aula Regular está intacta")
    print("="*80)
    
    file_path = "PCIELO-RESULTADOS-ICFES-MODELO-AULA-REGULAR-2025.xlsx"
    
    try:
        df = pd.read_excel(file_path)
        
        # Verificar que tiene datos
        if len(df) < 38:
            print(f"❌ FALLO: Archivo tiene {len(df)} filas, esperado al menos 38")
            return False
        
        # Verificar que fila 38 (índice 37) tiene datos 2024
        puntaje_2024 = df.iloc[37]['Puntaje Global']
        
        if puntaje_2024 <= 0:
            print(f"❌ FALLO: Puntaje global 2024 es {puntaje_2024}")
            return False
        
        print("✅ PASÓ: Archivo Aula Regular intacto")
        print(f"   - Total de filas: {len(df)}")
        print(f"   - Puntaje global 2024: {puntaje_2024:.2f}")
        return True
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return False

def main():
    """Ejecuta todos los tests."""
    print("\n" + "="*80)
    print("PRUEBAS DE VERIFICACIÓN - CAMBIOS APLICADOS")
    print("="*80)
    
    tests = [
        test_excel_flexible,
        test_pandas_indices,
        test_codigo_python,
        test_aula_regular_intacta
    ]
    
    resultados = []
    for test in tests:
        resultados.append(test())
    
    # Resumen
    print("\n" + "="*80)
    print("RESUMEN DE PRUEBAS")
    print("="*80)
    
    total = len(resultados)
    pasadas = sum(resultados)
    
    print(f"\nPruebas pasadas: {pasadas}/{total}")
    
    if all(resultados):
        print("\n✅ TODAS LAS PRUEBAS PASARON CORRECTAMENTE")
        print("\nLa aplicación está lista para publicar con:")
        print("  git add .")
        print("  git commit -m 'Actualizar Modelo Flexible con puntaje global 2024'")
        print("  git push origin main")
        return 0
    else:
        print("\n❌ ALGUNAS PRUEBAS FALLARON")
        print("Por favor, revisa los errores arriba")
        return 1

if __name__ == "__main__":
    sys.exit(main())

